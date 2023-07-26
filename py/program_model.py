import tkinter as tk
from tkinter import ttk, filedialog
import shutil
import os
import json
import openpyxl

class Program_model:

	def __init__(self):

		# 获取当前文件的目录路径
		self.current_directory = os.path.dirname(os.path.abspath(__file__))

		self.new_directory = self.current_directory.replace("\\py","")

		os.chdir(self.new_directory)	

		self.saved_modules = []	
		self.properties = {}
		self.module_data = ""
		self.module_dict = {}
		self.treeview_data = ""

		# 从文件中读取每一行的数据
		with open("Resource/Module/ModuleList.txt", "r") as file:
			self.module_list = file.readlines()		

	def matrix_path(self):
		file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("Excel Files", "*.xlsm")])
		file_dir, file_name = os.path.split(file_path)
		file_name_without_ext = os.path.splitext(file_name)[0]
		ext = os.path.splitext(file_name)[1]
		if ext == "xlsm":
			output_file = os.path.join(file_dir, file_name_without_ext + ".xlsx")
			shutil.copy(file_path, output_file)
		else: output_file = file_path
		if not output_file:
			return
		wb = openpyxl.load_workbook(output_file, data_only=True)
		base_found = False
		datasheets = {}
		for sheetname in wb.sheetnames:
			if sheetname == "BaseZone":
				base_found = True
			elif base_found:
				datasheets[sheetname] = wb[sheetname]
		return datasheets

	def lastRow(self, sheet, column):
		for cell in reversed(sheet[column]):
			if cell.value is not None:
				return cell.row

	def findCellAddress(self, sheet, string, col_num):
		for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_num, max_col=col_num):
			for cell in row:
				if cell.value == string:
					return cell.row

	def zoneArray(self, datasheets):
		zones = [key for key in datasheets]
		return zones

	def allStations(self, datasheets, zonelist):
		line_stations = {}
		for zone in zonelist:
			zone_stations = []
			sheet = datasheets[zone]
			firstST = self.findCellAddress(sheet, "Station Name", 1)
			lastST = self.lastRow(sheet, "A")
			for row in sheet.iter_rows(min_row=firstST, max_row=lastST, min_col=1, max_col=1):
				for cell in row:
					if cell.value != "Station Name" and cell.value != None:
						zone_stations.append(cell.value)
			line_stations[zone]=zone_stations
		return line_stations

	def allGates(self, datasheets, zonelist):
		line_gates = {}
		for zone in zonelist:
			zone_gates = []
			sheet = datasheets[zone]
			firstG = 7
			lastG = self.findCellAddress(sheet, "Station Name", 1) - 2
			print(f"firstST is {firstG} ,lastRow is {lastG}")
			for row in sheet.iter_rows(min_row=firstG, max_row=lastG, min_col=3, max_col=5):
				device_type = row[0].value
				device_name = row[2].value
				if device_type == "Gate":
					zone_gates.append(device_name)
			line_gates[zone] = zone_gates
		return line_gates

	def allRobots(self, datasheets, zonelist):
		line_robots = {}
		for zone in zonelist:
			zone_robots = {}
			sheet = datasheets[zone]
			firstRB = 7
			lastRB = self.findCellAddress(sheet, "Station Name", 1) - 2
			for row in sheet.iter_rows(min_row=firstRB, max_row=lastRB, min_col=3, max_col=5):
				device_type = row[0].value
				if device_type == "FANUC":				
					station_name = row[1].value
					if station_name not in zone_robots:
						zone_robots[station_name] = []
					device_name = row[2].value
					zone_robots[station_name].append(station_name + device_name)
			line_robots[zone] = zone_robots
		return line_robots

	def allDrives(self, datasheets, zonelist):
		line_drives = {}
		for zone in zonelist:
			zone_drives = {}
			sheet = datasheets[zone]
			firstDRV = 7
			lastDRV = self.findCellAddress(sheet, "Station Name", 1) - 2
			for row in sheet.iter_rows(min_row=firstDRV, max_row=lastDRV, min_col=3, max_col=5):
				device_type = row[0].value
				if device_type == "Drive":				
					station_name = row[1].value
					if station_name not in zone_drives:
						zone_drives[station_name] = []
					device_name = row[2].value
					zone_drives[station_name].append(station_name + device_name)
			line_drives[zone] = zone_drives
		return line_drives

	def allTooling(self, datasheets, zonelist):
		line_tools = {}
		for zone in zonelist:
			zone_tools = {}
			sheet = datasheets[zone]
			firstTL = 7
			lastTL = self.findCellAddress(sheet, "Station Name", 1) - 2
			for row in sheet.iter_rows(min_row=firstTL, max_row=lastTL, min_col=3, max_col=5):
				device_type = row[0].value
				if device_type == "VM":
					station_name = row[1].value
					if station_name not in zone_tools:
						zone_tools[station_name] = []
					device_name = row[2].value
					zone_tools[station_name].append(station_name + device_name)
			line_tools[zone] = zone_tools
		return line_tools

	def allCNV(self, datasheets, zonelist):
		line_CNV = {}
		for zone in zonelist:
			zone_tools = {}
			sheet = datasheets[zone]
			firstCNV = 7
			lastCNV = self.findCellAddress(sheet, "Station Name", 1) - 2
			for row in sheet.iter_rows(min_row=firstCNV, max_row=lastCNV, min_col=3, max_col=5):
				device_type = row[0].value
				if device_type == "CNV":
					station_name = row[1].value
					if station_name not in zone_CNV:
						zone_CNV[station_name] = []
					device_name = row[2].value
					zone_CNV[station_name].append(station_name + device_name)
			line_CNV[zone] = zone_CNV
		return line_CNV

	def allEstops(self, datasheets, zonelist):
		line_Estops = {}
		for zone in zonelist:
			zone_Estops = {}
			sheet = datasheets[zone]
			firstES= 7
			lastES = self.findCellAddress(sheet, "Station Name", 1) - 2
			for row in sheet.iter_rows(min_row=firstES, max_row=lastES, min_col=3, max_col=5):
				device_type = row[0].value
				station_name = row[1].value
				if device_type == "ESTOP" and station_name != None:
					if station_name not in zone_Estops:
						zone_Estops[station_name] = 0
					zone_Estops[station_name] += 1
			line_Estops[zone] = zone_Estops
		return line_Estops

	def allLCs(self, datasheets, zonelist):
		line_LCs = {}
		for zone in zonelist:
			zone_LCs = {}
			sheet = datasheets[zone]
			firstLC= 7
			lastLC = self.findCellAddress(sheet, "Station Name", 1) - 2
			for row in sheet.iter_rows(min_row=firstLC, max_row=lastLC, min_col=3, max_col=5):
				device_type = row[0].value
				station_name = row[1].value
				if device_type == "LC/SCN" and station_name != None:
					if station_name not in zone_LCs:
						zone_LCs[station_name] = 0
					zone_LCs[station_name] += 1
			line_LCs[zone] = zone_LCs
		return line_LCs

	def allOPs(self, datasheets, zonelist):
		line_OPs = {}
		for zone in zonelist:
			zone_OPs = {}
			sheet = datasheets[zone]
			firstOP= 7
			lastOP = self.findCellAddress(sheet, "Station Name", 1) - 2
			for row in sheet.iter_rows(min_row=firstOP, max_row=lastOP, min_col=3, max_col=5):
				device_type = row[0].value
				station_name = row[1].value
				if device_type == "Operator" and station_name != None:
					if station_name not in zone_OPs:
						zone_OPs[station_name] = 0
					zone_OPs[station_name] += 1
			line_OPs[zone] = zone_OPs
		return line_OPs

	def save_treeview_data(self, treeview, parent="", node_data_list=None):
		if node_data_list is None:
			node_data_list = []

		nodes = treeview.get_children(parent)
		for node in nodes:
			text = treeview.item(node, "text")
			values = treeview.item(node, "values")
			children_data = []
			
			# 递归保存子节点
			save_treeview_data(treeview, parent=node, node_data_list=children_data)
			
			node_data = {"text": text, "values": values, "children": children_data}
			node_data_list.append(node_data)

		return node_data_list	  

	def find_index(self, string, findstring, index):
		start_index = -1
		for i in range(index):
			start_index = string.find(findstring, start_index + 1)
			if start_index == -1:
				return -1
		return start_index

	def join_XIC(self, mode, line, Replace_string, itemlist, fullstring = "", index = 1):
		if mode == 1:
			conn = ""
		if mode == 2:
			conn = ","
		if fullstring == "":
			start_index = find_index(line, "XIC", index)
			end_index = find_index(line, ")", index) + 1
			str_XIC = line[start_index:end_index]
		else: 
			str_XIC = fullstring
		if Replace_string in line and str_XIC in line:
			new_XIC = conn.join(str_XIC.replace(Replace_string, item) for item in itemlist)
			new_line = line.replace(str_XIC, new_XIC)
		else: new_line = line
		return new_line

	def join_OTE_Branch(self, line, Replace_string, itemlist):
		start_index = line.find("[") + 1
		end_index = line.find("]") 
		str_OTE = line[start_index:end_index]
		new_OTE = ",".join(str_OTE.replace(Replace_string, item) for item in itemlist)
		return line.replace(str_OTE, new_OTE)

	def safety_StationHead(self, Station_name):
		with open("Resource/Program/SafetyTask/StationHead.txt", 'r') as file:
			lines = file.readlines()
		new_lines = [line.replace("Station_Name", Station_name) for line in lines]
		return new_lines		

	def safety_StationTags(self,Rblist):
		with open("Resource/Program/SafetyTask/StationTag.txt", 'r') as file:
			lines = file.readlines()
		new_lines = []
		for line in lines:
			if lines.index(line)==1:	
				for i in Range(len(Rblist)):
					new_line = line.replace("R01", str(i + 1))
					new_lines.append(new_line)
			else:
				new_lines.append(line)
		return new_lines		

	def safety_StationMain(self, Rblist,Toollist):
		with open("Resource/Program/SafetyTask/StationMain.txt", 'r') as file:
			lines = file.readlines()
		new_lines = []
		for line in lines:
			if lines.index(line)==7:	
				for robot in Rblist:
					new_line = line.replace("RB01", robot)
					new_lines.append(new_line)
			if lines.index(line) == 8:
				for tooling in Toollist:
					new_line = line.replace("TL01", tooling)
					new_lines.append(new_line)
			else:
				new_lines.append(line)
		return new_lines

	def safety_Mods(self,Station_Name, SIOlist):
		with open("Resource/Program/SafetyTask/StationMod.txt", 'r') as file:
			lines = file.readlines()
		new_lines = lines
		copied_lines = []
		for SIO in SIOlist:
			for line in lines[5:14]:
				line = line.replace("Module 1", "Module " + str(SIOlist.index(SIO) + 1))
				copied_lines.append(line.replace("SIO_Name", SIO))
		del new_lines[5:14]
		new_lines[5:5] = copied_lines
		lens = len(new_lines)
		count = len(SIOlist)
		for i in range(lens-4,lens-1):
			Mods = ["Mod" + str(i) for i in range(1, count+1)]
			new_XIC = join_XIC(new_lines[i], "Mod1", Mods)
			new_lines[i] = new_XIC
		new_lines = [line.replace("Station_Name", Station_Name) for line in new_lines]
		return new_lines

	def Estop(self, NumofEstops, Station_Name, Zone_Name):
		if NumofEstops > 0:
			with open("Resource/Program/SafetyTask/Estop.txt", 'r') as file:
				lines = file.readlines()
			lines = [line.replace("Station_Name", Station_Name) for line in lines]
			lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
			copied_lines = []
			for i in range(2, NumofEstops+1):
				for line in lines[5:15]:
					copied_lines.append(line.replace("E1", "E" + str(i)))
			insert_index = len(lines) - 1
			lines[insert_index:insert_index] = copied_lines
		else: lines = []
		return lines

	def LC(NumofLC, Station_Name, Zone_Name):
		if NumofLC > 0:
			with open("Resource/Program/SafetyTask/LC.txt", 'r') as file:
				lines = file.readlines()
			lines = [line.replace("Station_Name", Station_Name) for line in lines]
			lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
			copied_lines = []
			for i in range(NumofLC):
				for line in lines[13:30]:
					copied_lines.append(line.replace("LC1", "LC" + str(i + 1)))
			del lines[13:30]
			insert_index = len(lines) - 100
			lines[insert_index:insert_index] = copied_lines
		else: lines = []
		return lines

	def AccLC(self, NumofAccLC, Station_Name, Zone_Name):
		if NumofAccLC > 0:
			with open("Resource/Program/SafetyTask/AccLC.txt", 'r') as file:
				lines = file.readlines()
			lines = [line.replace("Station_Name", Station_Name) for line in lines]
			lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
		else: lines = []
		return lines

	def operator(self, NumofOP, Station_Name, Zone_Name):
		if NumofOP > 0:
			with open("Resource/Program/SafetyTask/Operator.txt", 'r') as file:
				lines = file.readlines()
			new_lines = [line.replace("Station_Name", Station_Name) for line in lines]
			new_lines = [line.replace("Zone_Name", Zone_Name) for line in new_lines]
		else: new_lines = []
		return new_lines

	def SPD(self, NumofSPD, Station_Name, Zone_Name):
		if NumofSPD > 0:
			with open("Resource/Program/SafetyTask/SPD.txt", 'r') as file:
				lines = file.readlines()
			lines = [line.replace("Station_Name", Station_Name) for line in lines]
			lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
			copied_lines = []
			for i in range(NumofSPD):
				for line in lines[5:10]:
					copied_lines.append(line.replace("SPD1", "SPD" + str(i + 1)))
			del lines[5:10]
			insert_index = len(lines) - 1
			lines[insert_index:insert_index] = copied_lines
		else: lines = []
		return lines

	def ICSD(self, hasICSD, Station_Name):
		if hasICSD:
			with open("Resource/Program/SafetyTask/StationICSD.txt", 'r') as file:
				lines = file.readlines()
			new_lines = [line.replace("Station_Name", Station_Name) for line in lines]
		else:
			new_lines = []
		return new_lines

	def safetyRobot(Self, Rblist, Station_Name, Zone_Name):
		if len(Rblist) > 0:
			with open("Resource/Program/SafetyTask/Robot.txt", 'r') as file:
				lines = file.readlines()
			lines = [line.replace("Station_Name", Station_Name) for line in lines]
			lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
			new_lines = []
			i = 1
			for robot in Rblist:
				for line in lines:
					if "RB01" in line:						
						new_lines.append(line.replace("RB01", "RB0" + str(i)))
					else:
						new_lines.append(line.replace("Robot_Name", robot))
				i+=1
			return new_lines

	def safetyTooling(self, Toollist, Station_Name, Zone_Name):
		if len(Toollist) > 0:
			with open("Resource/Program/SafetyTask/Tooling.txt", 'r') as file:
				lines = file.readlines()
			lines = [line.replace("Station_Name", Station_Name) for line in lines]
			lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
			new_lines = []
			i = 1
			for tooling in Toollist:
				tool_name = tooling.replace(Station_Name,"")
				for line in lines:
					if "TL1" in line:						
						new_lines.append(line.replace("TL1", tool_name))
					else:
						new_lines.append(line.replace("FX1", "FX" + str(i)))
				i+=1
			return new_lines

	def safety_StationSummary(self, Station_Name, Zone_Name):
		with open("Resource/Program/SafetyTask/StationSummary.txt", 'r') as file:
			lines = file.readlines()
		lines = [line.replace("Station_Name", Station_Name) for line in lines]
		lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
		return lines

	def safetyDrive(self, Drivelist, Station_Name, Zone_Name):
		if len(Drivelist) > 0:
			with open("Resource/Program/SafetyTask/Drive.txt", 'r') as file:
				lines = file.readlines()
			lines = [line.replace("Station_Name", Station_Name) for line in lines]
			lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
			new_lines = []
			i = 1
			for drive in Drivelist:
				drive_name = drive.replace(Station_Name,"")
				for line in lines:
					if "TT1" in line:						
						new_lines.append(line.replace("TT1", drive_name))
					else:
						new_lines.append(line.replace("Servo_Name", drive_name + "SRV1"))
				i+=1
			return new_lines

	def Gate(self, Gatelist, Zone_Name):
		if len(Gatelist) > 0:
			with open("Resource/Program/SafetyTask/Gate.txt", 'r') as file:
				lines = file.readlines()
			lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
			new_lines = []
			copied_lines = []
			for line in lines:	
				new_str = ""			
				if "[" in line:
					str_OTE = join_OTE(line, "Gate_Name", Gatelist)
					new_lines.append(str_OTE)								
				else: new_lines.append(line)
			for gate in Gatelist:
				for line in lines[25:33]:
					copied_lines.append(line.replace("Gate_Name", gate))
			del lines[25:33]
			insert_index = len(lines) - 1
			new_lines[insert_index:insert_index] = copied_lines
		else: new_lines = []
		return new_lines

	def safety_ZoneMain(self, Zone_Name):
		with open("Resource/Program/SafetyTask/ZoneMain.txt", 'r') as file:
			lines = file.readlines()
		lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
		return lines

	def safety_ZoneCommon(self, Zone_Name, Zone_index):
		with open("Resource/Program/SafetyTask/ZoneCommon.txt", 'r') as file:
			lines = file.readlines()
		lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
		if Zone_index > 0:
			del lines[4:6]
		return lines

	def ZoneSftyIn(self, Stationlist, Gatelist, Rblist, Drivelist, CNVlist, Zone_Name, Next_Zone):
		with open("Resource/Program/SafetyTask/ZoneSftyIn.txt", 'r') as file:
			lines = file.readlines()
		lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
		lines = [line.replace("Next_Zone", Next_Zone[0]) for line in lines]
		new_lines = []
		n = len(Rblist)
		mid = n // 2
		Rblist1 = Rblist[:mid + n % 2]
		Rblist2 = Rblist[mid + n % 2:]
		for line in lines:
			if "[" in line:
				line = join_XIC(2, line, "Gate_Name", Gatelist)
				line = join_XIC(2, line, "Robot_Name", Rblist)
				line = join_XIC(2, line, "Station_Name", Stationlist)
			elif "EStops1" in line or "Enbld1" in line:
				line = join_XIC(1, line, "Robot_Name", Rblist1, "XIC(Robot_NameSfty.Int.SftyIntEnbld)XIO(Robot_NameSfty.Int.SftyChanged)")
				line = join_XIC(1, line, "Robot_Name", Rblist1)
			elif "EStops2" in line or "Enbld2" in line:
				line = join_XIC(1, line, "Robot_Name", Rblist2, "XIC(Robot_NameSfty.Int.SftyIntEnbld)XIO(Robot_NameSfty.Int.SftyChanged)")
				line = join_XIC(1, line, "Robot_Name", Rblist2)
			else:
				line = join_XIC(1, line, "Station_Name", Stationlist)
				line = join_XIC(1, line, "Gate_Name", Gatelist)
				line = join_XIC(1, line, "Drive_Name", Drivelist)
				line = join_XIC(1, line, "Robot_Name", Rblist)
				line = join_XIC(1, line, "CNV_Name", CNVlist)
			if "Next_Zone" in line:
				line = join_XIC(1, line, Next_Zone[0], Next_Zone, "XIC(sNext_Zone.GateEStopsOk)")
				line = join_XIC(1, line, Next_Zone[0], Next_Zone, "XIC(sNext_Zone.AdjEStopsOk)")
				line = join_XIC(1, line, Next_Zone[0], Next_Zone, "XIC(sNext_Zone.OperEStopsOk)")
			new_lines.append(line)
		return new_lines

	def safety_ZoneSummary(self, Zone_Name):
		with open("Resource/Program/SafetyTask/ZoneSummary.txt", 'r') as file:
			lines = file.readlines()
		lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
		return lines

	def ZoneSftyOut(self, Zone_Name, Zone_index):
		with open("Resource/Program/SafetyTask/ZoneSftyOut.txt", 'r') as file:
			lines = file.readlines()
		lines = [line.replace("Zone_Name", Zone_Name) for line in lines]
		lines = [line.replace("Z1", "Z" + str(Zone_index + 1)) for line in lines]
		return lines

	def endprogram(self):
		with open("Resource/Program/EndProgram.txt", 'r') as file:
			lines = file.readlines()
		return lines





