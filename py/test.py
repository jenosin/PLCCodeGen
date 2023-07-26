import openpyxl
import os
from tkinter import ttk, filedialog
import shutil

def matrix_open():
	file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("Excel Files", "*.xlsm")])
	file_dir, file_name = os.path.split(file_path)
	file_name_without_ext = os.path.splitext(file_name)[0]
	ext = os.path.splitext(file_name)[1]
	if ext == "xlsm":
		output_file = os.path.join(file_dir, file_name_without_ext + ".xlsx")
		shutil.copy(file_path, output_file)
	else: output_file = file_path
	if not output_file:
		return
	wb = openpyxl.load_workbook(output_file, data_only=True)
	base_found = False
	datasheets = {}
	for sheetname in wb.sheetnames:
		if sheetname == "BaseZone":
			base_found = True
		elif base_found:
			datasheets[sheetname] = wb[sheetname]
	return datasheets

def lastRow(sheet, column):
	for cell in reversed(sheet[column]):
		if cell.value is not None:
			return cell.row

def findCellAddress(sheet, string, col_num):
	for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_num, max_col=col_num):
		for cell in row:
			if cell.value == string:
				return cell.row

def zoneArray(datasheets):
	zones = [key for key in datasheets]
	return zones

def allStations(datasheets, zonelist):
	line_stations = {}
	for zone in zonelist:
		zone_stations = []
		sheet = datasheets[zone]
		firstST = findCellAddress(sheet, "Station Name", 1)
		lastST = lastRow(sheet, "A")
		for row in sheet.iter_rows(min_row=firstST, max_row=lastST, min_col=1, max_col=1):
			if row[0].value != "Station Name" and row[0].value != None:
				zone_stations.append(row[0].value)
		line_stations[zone]=zone_stations
	return line_stations

def allOPs(datasheets, zonelist):
	line_OPs = {}
	for zone in zonelist:
		zone_OPs = {}
		sheet = datasheets[zone]
		firstOP= 7
		lastOP = findCellAddress(sheet, "Station Name", 1) - 2
		for row in sheet.iter_rows(min_row=firstOP, max_row=lastOP, min_col=3, max_col=5):
			device_type = row[0].value
			station_name = row[1].value
			if device_type == "Operator" and station_name != None:
				if station_name not in zone_OPs:
					zone_OPs[station_name] = 0
				zone_OPs[station_name] += 1
		line_OPs[zone] = zone_OPs
	return line_OPs


datasheets = matrix_open()
zones = zoneArray(datasheets)
AllOPs = allOPs(datasheets, zones)
print(AllOPs)
